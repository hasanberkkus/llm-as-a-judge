import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from src.report_generator import ReportGenerator


class ReportGeneratorTests(unittest.TestCase):
    def test_scoring_reports_and_columns_are_created(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            report_dir = Path(directory)
            with patch("src.report_generator.REPORTS_DIR", report_dir):
                generator = ReportGenerator()
                generator.generate(
                    dataset={
                        "expected_answers": [
                            {"question_id": 1, "expected_answer": "Referans"}
                        ]
                    },
                    results=[
                        {
                            "evaluation_mode": "scoring",
                            "question_id": 1,
                            "question_type": "calculation",
                            "question": "Soru?",
                            "correct_answer": "42",
                            "worker_answer": "42",
                            "verdict": "Correct",
                            "score": 10,
                            "reason": "Doğru.",
                        }
                    ],
                    worker_model="worker",
                    judge_model="judge",
                    dataset_name="text_1",
                )

            self.assertTrue(generator.json_path.exists())
            self.assertTrue(generator.txt_path.exists())
            self.assertTrue(generator.excel_path.exists())
            sheet = load_workbook(generator.excel_path).active
            headers = [cell.value for cell in sheet[1]]
            self.assertIn("Expected Answer", headers)
            self.assertIn("Judge Correct Answer", headers)
            self.assertIn("Score", headers)
            expected_answer_column = headers.index("Expected Answer") + 1
            self.assertEqual(sheet.cell(2, expected_answer_column).value, "Referans")
            self.assertIn(
                "Expected Answer : Referans",
                generator.txt_path.read_text(encoding="utf-8"),
            )


if __name__ == "__main__":
    unittest.main()
