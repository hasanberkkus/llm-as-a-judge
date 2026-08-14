from __future__ import annotations

import json
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from src.config import REPORTS_DIR


class ReportGenerator:
    """JSON, TXT ve Excel raporlarını oluşturan sınıf."""

    def __init__(self) -> None:
        REPORTS_DIR.mkdir(parents=True, exist_ok=True)

        self.json_path = REPORTS_DIR / "report.json"
        self.txt_path = REPORTS_DIR / "report.txt"
        self.excel_path = REPORTS_DIR / "LLM_Judge_Report.xlsx"

    def generate(
        self,
        dataset: dict[str, Any],
        results: list[dict[str, Any]],
        worker_model: str,
        judge_model: str,
        dataset_name: str,
    ) -> None:
        """Tüm raporları üretir."""
        if not results:
            raise ValueError(
                "Rapor oluşturmak için sonuç listesi boş olamaz."
            )

        evaluation_modes = {
            result["evaluation_mode"]
            for result in results
        }

        if len(evaluation_modes) != 1:
            raise ValueError(
                "Aynı rapordaki bütün sonuçlar aynı çalışma moduna "
                "sahip olmalıdır."
            )

        evaluation_mode = evaluation_modes.pop()

        if evaluation_mode not in {"scoring", "feedback"}:
            raise ValueError(
                f"Geçersiz çalışma modu: {evaluation_mode}"
            )

        self._generate_json(results)
        self._generate_txt(
            dataset=dataset,
            results=results,
            evaluation_mode=evaluation_mode,
            dataset_name=dataset_name,
        )
        self._generate_excel(
            dataset=dataset,
            results=results,
            worker_model=worker_model,
            judge_model=judge_model,
            evaluation_mode=evaluation_mode,
            dataset_name=dataset_name,
        )

        print(f"✔ JSON raporu oluşturuldu : {self.json_path}")
        print(f"✔ TXT raporu oluşturuldu  : {self.txt_path}")
        print(f"✔ Excel raporu oluşturuldu: {self.excel_path}")

    def _generate_json(
        self,
        results: list[dict[str, Any]],
    ) -> None:
        """Sonuçları JSON dosyasına değiştirmeden yazar."""
        with self.json_path.open("w", encoding="utf-8") as file:
            json.dump(
                results,
                file,
                ensure_ascii=False,
                indent=4,
            )

    def _generate_txt(
        self,
        dataset: dict[str, Any],
        results: list[dict[str, Any]],
        evaluation_mode: str,
        dataset_name: str,
    ) -> None:
        """Sonuçları çalışma moduna uygun TXT raporuna yazar."""
        expected_answers = self._get_expected_answers(dataset)
        dataset_label = self._get_dataset_label(dataset_name)

        with self.txt_path.open("w", encoding="utf-8") as file:
            file.write(f"Dataset : {dataset_label}\n\n")

            for result in results:
                expected_answer = expected_answers.get(
                    result["question_id"],
                    "Beklenen cevap bulunamadı.",
                )

                if evaluation_mode == "scoring":
                    self._write_scoring_txt_result(
                        file=file,
                        result=result,
                        expected_answer=expected_answer,
                    )
                else:
                    self._write_feedback_txt_result(
                        file=file,
                        result=result,
                        expected_answer=expected_answer,
                    )

    def _write_scoring_txt_result(
        self,
        file: Any,
        result: dict[str, Any],
        expected_answer: str,
    ) -> None:
        """Tek bir scoring sonucunu TXT raporuna yazar."""
        file.write(f"Soru {result['question_id']}\n")
        file.write(f"Soru : {result['question']}\n")
        file.write(f"Tür : {result['question_type']}\n")
        file.write("Çalışma Modu : Yalnızca puanlama\n")
        file.write(f"Expected Answer : {expected_answer}\n")
        file.write(
            f"Judge Correct Answer : {result['correct_answer']}\n"
        )
        file.write(f"Worker Answer : {result['worker_answer']}\n")
        file.write(f"Verdict : {result['verdict']}\n")
        file.write(f"Score : {result['score']}\n")
        file.write(f"Reason : {result['reason']}\n")
        file.write("-" * 80 + "\n")

    def _write_feedback_txt_result(
        self,
        file: Any,
        result: dict[str, Any],
        expected_answer: str,
    ) -> None:
        """Tek bir feedback sonucunu geçmişiyle TXT raporuna yazar."""
        attempts = self._get_attempts(result)
        initial_result = attempts[0]
        selected_attempt = result["selected_attempt"]
        initial_worker_answer = result.get(
            "initial_worker_answer",
            initial_result["worker_answer"],
        )
        initial_verdict = result.get(
            "initial_verdict",
            initial_result["verdict"],
        )
        initial_score = result.get(
            "initial_score",
            initial_result["score"],
        )
        initial_feedback = result.get(
            "initial_feedback",
            initial_result["feedback"],
        )
        attempt_count = result.get("attempt_count", len(attempts))
        selected_worker_answer = result.get(
            "selected_worker_answer",
            result["worker_answer"],
        )
        revision_applied = result.get(
            "revision_applied",
            len(attempts) > 1,
        )

        file.write(f"Soru {result['question_id']}\n")
        file.write(f"Soru : {result['question']}\n")
        file.write(f"Tür : {result['question_type']}\n")
        file.write("Çalışma Modu : Feedback ile düzeltme\n")
        file.write(f"Expected Answer : {expected_answer}\n")
        file.write(
            "Initial Worker Answer : "
            f"{initial_worker_answer}\n"
        )
        file.write(f"Initial Verdict : {initial_verdict}\n")
        file.write(f"Initial Score : {initial_score}\n")
        file.write(f"Initial Feedback : {initial_feedback}\n")
        file.write(f"Attempt Count : {attempt_count}\n")
        file.write(f"Selected Attempt : {selected_attempt}\n")
        file.write(
            f"Selected Worker Answer : {selected_worker_answer}\n"
        )
        file.write(f"Final Verdict : {result['verdict']}\n")
        file.write(f"Final Score : {result['score']}\n")
        file.write(f"Final Reason : {result['reason']}\n")
        file.write(
            "Revision Applied : "
            f"{'Evet' if revision_applied else 'Hayır'}\n"
        )
        file.write("\nATTEMPT HISTORY\n")

        for attempt_result in attempts:
            attempt_number = attempt_result.get("attempt", 1)
            file.write(f"\nAttempt Number : {attempt_number}\n")
            file.write(
                f"Worker Answer : {attempt_result['worker_answer']}\n"
            )
            file.write(
                "Judge Correct Answer : "
                f"{attempt_result['correct_answer']}\n"
            )
            file.write(f"Verdict : {attempt_result['verdict']}\n")
            file.write(f"Score : {attempt_result['score']}\n")
            file.write(f"Reason : {attempt_result['reason']}\n")
            file.write(f"Feedback : {attempt_result['feedback']}\n")
            file.write(
                "Selected : "
                f"{'Evet' if attempt_number == selected_attempt else 'Hayır'}\n"
            )

        file.write("-" * 80 + "\n")

    def _generate_excel(
        self,
        dataset: dict[str, Any],
        results: list[dict[str, Any]],
        worker_model: str,
        judge_model: str,
        evaluation_mode: str,
        dataset_name: str,
    ) -> None:
        """Çalışma moduna uygun Excel raporunu oluşturur."""
        workbook = Workbook()
        current_time = datetime.now()
        experiment_id = current_time.strftime("EXP_%Y%m%d_%H%M%S")
        experiment_date = current_time.strftime("%d.%m.%Y %H:%M")

        self._create_experiment_sheet(
            workbook=workbook,
            dataset=dataset,
            results=results,
            worker_model=worker_model,
            judge_model=judge_model,
            evaluation_mode=evaluation_mode,
            experiment_id=experiment_id,
            experiment_date=experiment_date,
            dataset_name=dataset_name,
        )

        if evaluation_mode == "feedback":
            self._create_attempt_history_sheet(
                workbook=workbook,
                results=results,
            )

        workbook.save(self.excel_path)

    def _create_experiment_sheet(
        self,
        workbook: Workbook,
        dataset: dict[str, Any],
        results: list[dict[str, Any]],
        worker_model: str,
        judge_model: str,
        evaluation_mode: str,
        experiment_id: str,
        experiment_date: str,
        dataset_name: str,
    ) -> None:
        """Deney sonuçlarının bulunduğu ana sayfayı oluşturur."""
        worksheet = workbook.active
        worksheet.title = "Experiment Results"
        expected_answers = self._get_expected_answers(dataset)

        if evaluation_mode == "scoring":
            headers = [
                "Experiment ID",
                "Date",
                "Evaluation Mode",
                "Dataset",
                "Worker Model",
                "Judge Model",
                "Question ID",
                "Question Type",
                "Question",
                "Expected Answer",
                "Judge Correct Answer",
                "Worker Answer",
                "Verdict",
                "Score",
                "Reason",
            ]
        else:
            headers = [
                "Experiment ID",
                "Date",
                "Evaluation Mode",
                "Dataset",
                "Worker Model",
                "Judge Model",
                "Question ID",
                "Question Type",
                "Question",
                "Expected Answer",
                "Judge Correct Answer",
                "Initial Worker Answer",
                "Initial Verdict",
                "Initial Score",
                "Initial Feedback",
                "Selected Worker Answer",
                "Selected Attempt",
                "Attempt Count",
                "Revision Applied",
                "Final Verdict",
                "Final Score",
                "Final Reason",
            ]

        self._write_headers(worksheet, headers)
        dataset_label = self._get_dataset_label(dataset_name)

        for index, result in enumerate(results, start=2):
            expected_answer = expected_answers.get(
                result["question_id"],
                "Beklenen cevap bulunamadı.",
            )

            if evaluation_mode == "scoring":
                values = [
                    experiment_id,
                    experiment_date,
                    "Yalnızca puanlama",
                    dataset_label,
                    worker_model,
                    judge_model,
                    result["question_id"],
                    result["question_type"],
                    result["question"],
                    expected_answer,
                    result["correct_answer"],
                    result["worker_answer"],
                    result["verdict"],
                    result["score"],
                    result["reason"],
                ]
                verdict_column = 13
            else:
                attempts = self._get_attempts(result)
                initial_result = attempts[0]
                initial_worker_answer = result.get(
                    "initial_worker_answer",
                    initial_result["worker_answer"],
                )
                initial_verdict = result.get(
                    "initial_verdict",
                    initial_result["verdict"],
                )
                initial_score = result.get(
                    "initial_score",
                    initial_result["score"],
                )
                initial_feedback = result.get(
                    "initial_feedback",
                    initial_result["feedback"],
                )
                attempt_count = result.get(
                    "attempt_count",
                    len(attempts),
                )
                selected_worker_answer = result.get(
                    "selected_worker_answer",
                    result["worker_answer"],
                )
                revision_applied = result.get(
                    "revision_applied",
                    len(attempts) > 1,
                )
                values = [
                    experiment_id,
                    experiment_date,
                    "Feedback ile düzeltme",
                    dataset_label,
                    worker_model,
                    judge_model,
                    result["question_id"],
                    result["question_type"],
                    result["question"],
                    expected_answer,
                    result["correct_answer"],
                    initial_worker_answer,
                    initial_verdict,
                    initial_score,
                    initial_feedback,
                    selected_worker_answer,
                    result["selected_attempt"],
                    attempt_count,
                    "Evet" if revision_applied else "Hayır",
                    result["verdict"],
                    result["score"],
                    result["reason"],
                ]
                verdict_column = 20

            self._write_row(
                worksheet=worksheet,
                row=index,
                values=values,
                alternate=index % 2 == 1,
            )
            self._color_verdict(
                worksheet=worksheet,
                row=index,
                column=verdict_column,
                verdict=result["verdict"],
            )

        self._finish_sheet(worksheet)

    def _create_attempt_history_sheet(
        self,
        workbook: Workbook,
        results: list[dict[str, Any]],
    ) -> None:
        """Feedback modundaki bütün attempt kayıtlarını yazar."""
        worksheet = workbook.create_sheet("Attempt History")
        headers = [
            "Question ID",
            "Attempt Number",
            "Worker Answer",
            "Judge Correct Answer",
            "Verdict",
            "Score",
            "Reason",
            "Feedback",
            "Selected",
        ]
        self._write_headers(worksheet, headers)

        row = 2

        for result_index, result in enumerate(results):
            for attempt_result in self._get_attempts(result):
                attempt_number = attempt_result.get("attempt", 1)
                values = [
                    result["question_id"],
                    attempt_number,
                    attempt_result["worker_answer"],
                    attempt_result["correct_answer"],
                    attempt_result["verdict"],
                    attempt_result["score"],
                    attempt_result["reason"],
                    attempt_result["feedback"],
                    (
                        "Evet"
                        if attempt_number == result["selected_attempt"]
                        else "Hayır"
                    ),
                ]
                self._write_row(
                    worksheet=worksheet,
                    row=row,
                    values=values,
                    alternate=result_index % 2 == 1,
                )
                self._color_verdict(
                    worksheet=worksheet,
                    row=row,
                    column=5,
                    verdict=attempt_result["verdict"],
                )
                row += 1

        self._finish_sheet(worksheet)

    @staticmethod
    def _get_expected_answers(
        dataset: dict[str, Any],
    ) -> dict[Any, str]:
        """Beklenen cevapları soru kimliğine göre eşler."""
        return {
            answer["question_id"]: answer["expected_answer"]
            for answer in dataset["expected_answers"]
        }

    @staticmethod
    def _get_dataset_label(dataset_name: str) -> str:
        """Dataset klasör adını raporda gösterilecek etikete çevirir."""
        dataset_labels = {
            "text_1": "Text 1",
            "text_2": "Text 2",
        }

        try:
            return dataset_labels[dataset_name]
        except KeyError as error:
            raise ValueError(
                f"Geçersiz dataset adı: {dataset_name}"
            ) from error

    @staticmethod
    def _get_attempts(
        result: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Sonuçtaki attempt geçmişini döndürür."""
        return result.get(
            "attempt_history",
            result.get("attempts", [result]),
        )

    @staticmethod
    def _write_headers(worksheet: Any, headers: list[str]) -> None:
        """Sayfanın başlık satırını biçimlendirir."""
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        border = ReportGenerator._create_border()

        for column, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=column, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        worksheet.row_dimensions[1].height = 32

    @staticmethod
    def _write_row(
        worksheet: Any,
        row: int,
        values: list[Any],
        alternate: bool,
    ) -> None:
        """Bir veri satırını ortak stille yazar."""
        fill = PatternFill(
            fill_type="solid",
            fgColor="EAF3FF" if alternate else "FFFFFF",
        )
        alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        )
        border = ReportGenerator._create_border()

        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row, column=column, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = alignment

        worksheet.row_dimensions[row].height = 100

    @staticmethod
    def _color_verdict(
        worksheet: Any,
        row: int,
        column: int,
        verdict: str,
    ) -> None:
        """Verdict hücresini sonucuna göre renklendirir."""
        verdict_colors = {
            "Correct": "C6EFCE",
            "Partially Correct": "FFF2CC",
            "Incorrect": "F4CCCC",
        }
        worksheet.cell(row=row, column=column).fill = PatternFill(
            fill_type="solid",
            fgColor=verdict_colors.get(verdict, "FFFFFF"),
        )

    @staticmethod
    def _create_border() -> Border:
        """Rapor hücreleri için ince kenarlık oluşturur."""
        thin_side = Side(style="thin", color="D9D9D9")
        return Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

    @staticmethod
    def _finish_sheet(worksheet: Any) -> None:
        """Filtre, sabitleme ve sütun genişliklerini uygular."""
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(
                (
                    len(str(cell.value))
                    for cell in column_cells
                    if cell.value is not None
                ),
                default=0,
            )
            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 3, 12),
                50,
            )
